import re
import io
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import streamlit as st

st.set_page_config(
    page_title="PDF → Excel Converter",
    page_icon="📊",
    layout="centered"
)

# ─── CSS (theme-aware — no hardcoded background) ────────────────────
st.markdown("""
<style>
.stat-label { font-size: 12px; margin-bottom: 4px; }
.stat-value { font-size: 14px; font-weight: 700; }
</style>
""", unsafe_allow_html=True)


# ─── Shared Helpers ──────────────────────────────────────────────────
def clean_num(val):
    if not val or str(val).strip() == '':
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except:
        return None

def clean_text(val):
    if not val:
        return ''
    return re.sub(r'[\uf700-\uf7ff]', '', str(val)).strip()


# ─── PDF Type Detection ──────────────────────────────────────────────
def detect_pdf_type(pdf_bytes):
    """Returns 'statement', 'matrix_trial_balance', or 'trial_balance'."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = (pdf.pages[0].extract_text() or '').upper()
            if 'STATEMENT OF ACCOUNT' in text:
                return 'statement'
            if 'MATRIX TRIAL BALANCE' in text:
                return 'matrix_trial_balance'
    except:
        pass
    return 'trial_balance'


# ─── Matrix Trial Balance ────────────────────────────────────────────
def detect_info_matrix(pdf_bytes):
    """Returns (company_name, report_date) from Matrix Trial Balance PDF.
    Line 0 format: 'Company Name DD/MM/YY'
    """
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            if lines:
                # First non-empty line: "Katathani Phuket Beach Resort 19/03/26"
                first = lines[0]
                m = re.search(r'(\d{2}/\d{2}/\d{2})\s*$', first)
                if m:
                    date    = m.group(1)
                    company = first[:m.start()].strip()
                else:
                    company = first
                    date    = ''
                return company, date
    except:
        pass
    return '', ''

def extract_matrix_rows(pdf_bytes):
    """Extract all data rows from a New Matrix Trial Balance PDF."""
    all_rows = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            rows_by_y = defaultdict(list)
            for w in words:
                rows_by_y[round(w['top'] / 2) * 2].append(w)

            # ── Detect column x1 anchors from "Debit/Credit" sub-header ──
            col_x1s = None
            header_sub_y = 0
            for y in sorted(rows_by_y.keys()):
                rw = sorted(rows_by_y[y], key=lambda w: w['x0'])
                dc_count = sum(1 for w in rw if w['text'] in ('Debit', 'Credit'))
                if dc_count >= 3:
                    col_x1s = sorted(w['x1'] for w in rw if w['text'] in ('Debit', 'Credit'))
                    header_sub_y = y
                    break

            if not col_x1s or len(col_x1s) < 8:
                continue

            # ── Parse each data row ──
            for y in sorted(rows_by_y.keys()):
                if y <= header_sub_y:
                    continue

                ws = sorted(rows_by_y[y], key=lambda w: w['x0'])
                if not ws:
                    continue
                # Must start with a pure-digit Trn Code at far left
                if not re.match(r'^\d+$', ws[0]['text']) or ws[0]['x0'] > 5:
                    continue

                trn_code = ws[0]['text']

                # ── Merge "- number" into a single negative token ──
                merged = []
                i = 1  # skip trn_code already recorded
                while i < len(ws):
                    w = ws[i]
                    if (w['text'] == '-' and i + 1 < len(ws) and
                            ws[i + 1]['x0'] - w['x1'] < 5 and
                            re.match(r'^[\d,]+', ws[i + 1]['text'])):
                        merged.append({'x0': w['x0'], 'x1': ws[i + 1]['x1'],
                                       'text': '-' + ws[i + 1]['text']})
                        i += 2
                    else:
                        merged.append(w)
                        i += 1

                # ── Assign words to fields using x1-based column detection ──
                desc_parts = []
                net_amt    = ''
                values     = {ci: '' for ci in range(len(col_x1s))}
                net_rev    = ''

                for w in merged:
                    x0, x1, t = w['x0'], w['x1'], w['text']
                    is_num = bool(re.match(r'^-?[\d,]+\.?\d*$', t))

                    if x1 < col_x1s[0] - 30:       # left zone: description or net_amt
                        if is_num:
                            net_amt = t
                        elif x0 >= 20:
                            desc_parts.append(t)
                    elif x1 > col_x1s[-1] + 5:     # right of last ledger col = Net Revenue
                        net_rev = t
                    else:
                        best_col = min(range(len(col_x1s)),
                                       key=lambda ci: abs(x1 - col_x1s[ci]))
                        if abs(x1 - col_x1s[best_col]) < 30:
                            values[best_col] = t

                all_rows.append({
                    'trn_code': trn_code,
                    'desc':     ' '.join(desc_parts),
                    'net_amt':  net_amt,
                    'dep_dr':   values.get(0, ''), 'dep_cr':   values.get(1, ''),
                    'guest_dr': values.get(2, ''), 'guest_cr': values.get(3, ''),
                    'pkg_dr':   values.get(4, ''), 'pkg_cr':   values.get(5, ''),
                    'ar_dr':    values.get(6, ''), 'ar_cr':    values.get(7, ''),
                    'int_db':   values.get(8, ''),
                    'net_rev':  net_rev,
                })

    return all_rows

def _matrix_words_to_vals(words, col_x1s):
    """Map a row of pdfplumber words → list of 11 floats/None:
    [net_amt, dep_dr, dep_cr, guest_dr, guest_cr, pkg_dr, pkg_cr, ar_dr, ar_cr, int_db, net_rev]
    Handles '- 12,345.67' negative sign pairs.
    """
    vals = [None] * 11
    ws = sorted(words, key=lambda w: w['x0'])
    # Merge "- number" pairs
    merged = []
    i = 0
    while i < len(ws):
        w = ws[i]
        if (w['text'] == '-' and i + 1 < len(ws) and
                ws[i+1]['x0'] - w['x1'] < 5 and
                re.match(r'^[\d,]+', ws[i+1]['text'])):
            merged.append({'x0': w['x0'], 'x1': ws[i+1]['x1'],
                           'text': '-' + ws[i+1]['text']})
            i += 2
        else:
            merged.append(w)
            i += 1
    for w in merged:
        x1, t = w['x1'], w['text']
        if not re.match(r'^-?[\d,]+\.?\d*$', t):
            continue
        v = clean_num(t)
        if x1 < col_x1s[0] - 30:         # Net Amount
            vals[0] = v
        elif x1 > col_x1s[-1] + 5:       # Net Revenue
            vals[10] = v
        else:
            best = min(range(len(col_x1s)), key=lambda ci: abs(x1 - col_x1s[ci]))
            if abs(x1 - col_x1s[best]) < 35:
                vals[best + 1] = v        # offset: vals[0]=net_amt, vals[1]=dep_dr, …
    return vals

def _extract_matrix_special(pdf_bytes):
    """Extract Balance-From, Running-Totals-BF, PDF-Total, Running-Totals-CF,
    and Final-Balance rows.  Returns a dict with keys:
      bf_date, bf_vals, rt_bf_vals, total_vals, rt_cf_vals, fb_vals
    Each *_vals is a list of 11 floats/None.
    """
    out = {}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            # ── Get col_x1s from first page ──
            page0 = pdf.pages[0]
            rby0 = defaultdict(list)
            for w in page0.extract_words():
                rby0[round(w['top'] / 2) * 2].append(w)
            col_x1s = None
            for y in sorted(rby0.keys()):
                rw = rby0[y]
                dc = sum(1 for w in rw if w['text'] in ('Debit', 'Credit'))
                if dc >= 3:
                    col_x1s = sorted(w['x1'] for w in rw if w['text'] in ('Debit', 'Credit'))
                    break
            if not col_x1s:
                return out

            # ── First page: Balance From + Running Totals BF ──
            for y in sorted(rby0.keys()):
                rw = sorted(rby0[y], key=lambda w: w['x0'])
                texts = [w['text'] for w in rw]
                if 'Balance' in texts and 'From' in texts:
                    for w in rw:
                        if re.match(r'^\d{2}/\d{2}/\d{2}$', w['text']):
                            out['bf_date'] = w['text']
                    out['bf_vals'] = _matrix_words_to_vals(rw, col_x1s)
                elif ('bf_vals' in out and 'rt_bf_vals' not in out and
                      not any(w['text'] in ('Debit', 'Credit', 'Balance', 'From') for w in rw)):
                    nums = [w for w in rw if re.match(r'^-?[\d,]+\.?\d*$', w['text'])]
                    if len(nums) >= 5:
                        out['rt_bf_vals'] = _matrix_words_to_vals(rw, col_x1s)

            # ── Last page: Total + Running Totals CF + Final Balance ──
            last = pdf.pages[-1]
            rby_last = defaultdict(list)
            for w in last.extract_words():
                rby_last[round(w['top'] / 2) * 2].append(w)

            total_y = None
            post_count = 0
            for y in sorted(rby_last.keys()):
                rw = sorted(rby_last[y], key=lambda w: w['x0'])
                texts = [w['text'] for w in rw]
                if 'Total' in texts and total_y is None:
                    total_y = y
                    out['total_vals'] = _matrix_words_to_vals(rw, col_x1s)
                elif total_y and y > total_y:
                    nums = [w for w in rw if re.match(r'^-?[\d,]+\.?\d*$', w['text'])
                            or w['text'] == '-']
                    if len(nums) >= 3:
                        if post_count == 0:
                            out['rt_cf_vals'] = _matrix_words_to_vals(rw, col_x1s)
                        elif post_count == 1:
                            out['fb_vals'] = _matrix_words_to_vals(rw, col_x1s)
                        post_count += 1
    except:
        pass
    return out

def convert_matrix_trial_balance(pdf_bytes, company_name, report_date):
    rows    = extract_matrix_rows(pdf_bytes)
    special = _extract_matrix_special(pdf_bytes)

    if not rows:
        raise ValueError("ไม่พบข้อมูลใน PDF — กรุณาตรวจสอบว่าเป็น Matrix Trial Balance ที่รองรับ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrix Trial Balance"

    white_f  = Font(name='Arial', bold=True, size=9,  color="FFFFFF")
    data_f   = Font(name='Arial', size=9)
    bold_f   = Font(name='Arial', bold=True, size=9)
    title_f  = Font(name='Arial', bold=True, size=13, color="1F4E79")
    sub_f    = Font(name='Arial', bold=True, size=10, color="2E75B6")
    c_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    r_align  = Alignment(horizontal='right',  vertical='center')
    thin     = Side(style='thin',   color='BFBFBF')
    medium   = Side(style='medium', color='2E75B6')
    t_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    sub_fill = PatternFill("solid", fgColor="2E75B6")
    tot_fill = PatternFill("solid", fgColor="D6E4F0")
    bf_fill  = PatternFill("solid", fgColor="EAF0FB")   # Balance-From rows
    rt_fill  = PatternFill("solid", fgColor="F5F5F5")   # Running-Totals rows
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    num_fmt  = '#,##0.00;[Red](#,##0.00);"-"'
    num_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]   # C..M

    def write_num_row(er, label, vals_11, fill, font, bold_border=False):
        """Write one special row: A=label(merged A:B), C..M = vals_11[0..10]."""
        bdr = Border(left=thin, right=thin,
                     top=(medium if bold_border else thin),
                     bottom=(medium if bold_border else thin))
        ws.merge_cells(f'A{er}:B{er}')
        c = ws.cell(row=er, column=1, value=label)
        c.font = font; c.fill = fill; c.alignment = l_align; c.border = bdr
        ws.cell(row=er, column=2).fill = fill
        ws.cell(row=er, column=2).border = bdr
        for col_idx, v in zip(num_cols, vals_11):
            c = ws.cell(row=er, column=col_idx, value=v)
            c.font = font; c.fill = fill; c.alignment = r_align; c.border = bdr
            if v is not None: c.number_format = num_fmt
        ws.row_dimensions[er].height = 16

    def write_balance_row(er, label, vals_11, fill, font):
        """Write a Balance-From / Balance-As-At row where each ledger has ONE
        net value (not split Debit/Credit).  Pairs D:E, F:G, H:I, J:K are
        merged and the value is centred inside the merged cell.
        vals_11 layout: [net_amt, dep, None, guest, None, pkg, None, ar, None, int_db, net_rev]
        """
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.merge_cells(f'A{er}:B{er}')
        c = ws.cell(row=er, column=1, value=label)
        c.font = font; c.fill = fill; c.alignment = l_align; c.border = bdr
        ws.cell(row=er, column=2).fill = fill
        ws.cell(row=er, column=2).border = bdr

        # C: Net Amount
        c = ws.cell(row=er, column=3, value=vals_11[0])
        c.font = font; c.fill = fill; c.alignment = r_align; c.border = bdr
        if vals_11[0] is not None: c.number_format = num_fmt

        # D:E → Deposit (single net value)
        # F:G → Guest,  H:I → Package,  J:K → A/R
        pairs = [
            (4, 5,  vals_11[1]),   # Deposit  cols D:E
            (6, 7,  vals_11[3]),   # Guest    cols F:G
            (8, 9,  vals_11[5]),   # Package  cols H:I
            (10, 11, vals_11[7]),  # A/R      cols J:K
        ]
        for col_start, col_end, v in pairs:
            ws.merge_cells(start_row=er, start_column=col_start,
                           end_row=er,   end_column=col_end)
            c = ws.cell(row=er, column=col_start, value=v)
            c.font = font; c.fill = fill; c.alignment = r_align; c.border = bdr
            if v is not None: c.number_format = num_fmt

        # L: Internal DB,  M: Net Revenue
        for col_idx, v in [(12, vals_11[9]), (13, vals_11[10])]:
            c = ws.cell(row=er, column=col_idx, value=v)
            c.font = font; c.fill = fill; c.alignment = r_align; c.border = bdr
            if v is not None: c.number_format = num_fmt

        ws.row_dimensions[er].height = 16

    # ── Rows 1-3: title / subtitle / spacer ──
    ws.merge_cells('A1:M1')
    ws['A1'] = company_name
    ws['A1'].font = title_f; ws['A1'].alignment = c_align
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:K2')
    ws['A2'] = 'New Matrix Trial Balance Report'
    ws['A2'].font = sub_f; ws['A2'].alignment = l_align
    ws['L2'] = 'Date:'; ws['L2'].font = Font(name='Arial', bold=True, size=9); ws['L2'].alignment = r_align
    ws['M2'] = report_date; ws['M2'].font = Font(name='Arial', size=9); ws['M2'].alignment = c_align
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 5

    # ── Rows 4-5: 2-level column headers ──
    for start, end, label in [
        ('A4','A5','Trn. Code'), ('B4','B5','Description'), ('C4','C5','Net Amount'),
        ('D4','E4','Deposit Ledger'), ('F4','G4','Guest Ledger'),
        ('H4','I4','Package Ledger'), ('J4','K4','A/R Ledger'),
        ('L4','L5','Internal DB'),   ('M4','M5','Net Revenue'),
    ]:
        if start != end: ws.merge_cells(f'{start}:{end}')
        c = ws[start]; c.value = label
        c.font = white_f; c.fill = hdr_fill; c.alignment = c_align; c.border = t_border

    for ref, lbl in [('D5','Debit'),('E5','Credit'),('F5','Debit'),('G5','Credit'),
                     ('H5','Debit'),('I5','Credit'),('J5','Debit'),('K5','Credit')]:
        c = ws[ref]; c.value = lbl
        c.font = white_f; c.fill = sub_fill; c.alignment = c_align; c.border = t_border
    for ref in ['A5','B5','C5','L5','M5']:
        ws[ref].fill = hdr_fill; ws[ref].border = t_border
    ws.row_dimensions[4].height = 22; ws.row_dimensions[5].height = 18

    cur_row = 6  # next Excel row to write

    # ── Balance From row (from PDF header) ──
    if 'bf_vals' in special:
        bf_label = f"Balance From  {special.get('bf_date','')}"
        write_balance_row(cur_row, bf_label, special['bf_vals'], bf_fill, bold_f)
        cur_row += 1

    # ── Running Totals BF row ──
    if 'rt_bf_vals' in special:
        write_num_row(cur_row, 'Running Totals B/F', special['rt_bf_vals'], rt_fill, data_f)
        cur_row += 1

    # ── Data rows ──
    data_start = cur_row
    for i, row in enumerate(rows):
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            row['trn_code'], row['desc'],
            clean_num(row['net_amt']),
            clean_num(row['dep_dr']),   clean_num(row['dep_cr']),
            clean_num(row['guest_dr']), clean_num(row['guest_cr']),
            clean_num(row['pkg_dr']),   clean_num(row['pkg_cr']),
            clean_num(row['ar_dr']),    clean_num(row['ar_cr']),
            clean_num(row['int_db']),   clean_num(row['net_rev']),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=cur_row, column=col, value=val)
            c.font = data_f; c.fill = fill; c.border = t_border
            if col == 1:   c.alignment = c_align
            elif col == 2: c.alignment = l_align
            else:
                c.alignment = r_align
                if val is not None: c.number_format = num_fmt
        ws.row_dimensions[cur_row].height = 14
        cur_row += 1

    # ── Total row — use PDF values (exact, no floating-point drift) ──
    total_vals = special.get('total_vals', [None] * 11)
    write_num_row(cur_row, 'Total', total_vals, tot_fill, bold_f, bold_border=True)
    cur_row += 1

    # ── Running Totals CF + Final Balance rows ──
    if 'rt_cf_vals' in special:
        write_num_row(cur_row, 'Running Totals C/F', special['rt_cf_vals'], rt_fill, data_f)
        cur_row += 1
    if 'fb_vals' in special:
        fb_label = f"Balance As At  {report_date}"
        write_balance_row(cur_row, fb_label, special['fb_vals'], bf_fill, bold_f)
        cur_row += 1

    # ── Column widths & sheet settings ──
    for i, cw in enumerate([9, 32, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = cw

    ws.freeze_panes = f'C{data_start}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1

    # Total Net Amount from PDF (index 0 of total_vals)
    total_net = total_vals[0] if total_vals and total_vals[0] is not None else 0.0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(rows), total_net


# ─── Trial Balance ───────────────────────────────────────────────────
def detect_company_tb(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            return lines[0] if lines else ''
    except:
        return ''

def convert_trial_balance(pdf_bytes, company_name):
    all_rows = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or row[0] in ['เลขที่บัญชี', None]:
                        continue
                    if row[2] in ['ยอดยกมา', 'เดบิต']:
                        continue
                    if not row[0] or not str(row[0]).strip():
                        continue
                    all_rows.append(row)

    if not all_rows:
        raise ValueError("ไม่พบข้อมูลตารางใน PDF — กรุณาตรวจสอบว่าเป็น PDF งบทดลองที่รองรับ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "งบทดลอง"

    white_f  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    data_f   = Font(name='Arial', size=10)
    total_f  = Font(name='Arial', bold=True, size=10)
    c_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align  = Alignment(horizontal='left',   vertical='center')
    r_align  = Alignment(horizontal='right',  vertical='center')
    thin     = Side(style='thin',   color='BFBFBF')
    medium   = Side(style='medium', color='2E75B6')
    t_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    sub_fill = PatternFill("solid", fgColor="2E75B6")
    tot_fill = PatternFill("solid", fgColor="D6E4F0")
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    num_fmt  = '#,##0.00;[Red](#,##0.00);"-"'

    ws.merge_cells('A1:H1')
    ws['A1'] = company_name
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color="1F4E79")
    ws['A1'].alignment = c_align

    ws.merge_cells('A2:H2')
    ws['A2'] = 'รายงานงบทดลอง'
    ws['A2'].font = Font(name='Arial', bold=True, size=12, color="2E75B6")
    ws['A2'].alignment = c_align

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8

    for cells in ['A4:A5', 'B4:B5', 'C4:D4', 'E4:F4', 'G4:H4']:
        ws.merge_cells(cells)
    top_labels = ['เลขที่บัญชี', 'ชื่อบัญชี', 'ยอดยกมา', '', 'ยอดสะสมประจำงวด', '', 'ยอดยกไป', '']
    sub_labels = ['', '', 'เดบิต', 'เครดิต', 'เดบิต', 'เครดิต', 'เดบิต', 'เครดิต']
    for col, val in enumerate(top_labels, 1):
        c = ws.cell(row=4, column=col, value=val or None)
        c.font = white_f; c.fill = hdr_fill; c.alignment = c_align; c.border = t_border
    for col, val in enumerate(sub_labels, 1):
        c = ws.cell(row=5, column=col, value=val or None)
        c.font = white_f; c.fill = sub_fill; c.alignment = c_align; c.border = t_border
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 20

    totals = [0.0] * 6
    for i, row in enumerate(all_rows):
        er   = 6 + i
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [clean_text(row[0]), clean_text(row[1]),
                clean_num(row[2]), clean_num(row[3]),
                clean_num(row[4]), clean_num(row[5]),
                clean_num(row[6]), clean_num(row[7])]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=er, column=col, value=val)
            c.font = data_f; c.fill = fill; c.border = t_border
            c.alignment = c_align if col == 1 else (l_align if col == 2 else r_align)
            if col > 2 and val is not None:
                c.number_format = num_fmt
        for j, v in enumerate(vals[2:]):
            if v: totals[j] += v

    tr = 6 + len(all_rows)
    for col in range(1, 9):
        c = ws.cell(row=tr, column=col)
        c.fill = tot_fill; c.font = total_f
        c.border = Border(left=thin, right=thin, top=medium, bottom=medium)
        c.alignment = c_align if col == 1 else r_align
    ws.cell(row=tr, column=1).value = 'รวม'
    for col_idx, total in zip([3, 4, 5, 6, 7, 8], totals):
        ws.cell(row=tr, column=col_idx).value = total
        ws.cell(row=tr, column=col_idx).number_format = num_fmt
    ws.row_dimensions[tr].height = 22

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 55
    for col_letter in 'CDEFGH':
        ws.column_dimensions[col_letter].width = 18
    ws.freeze_panes = 'C6'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(all_rows), totals


# ─── Statement of Account ────────────────────────────────────────────
def detect_company_soa(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            name_parts = []
            for l in lines[:6]:
                if 'STATEMENT OF ACCOUNT' in l.upper(): continue
                if 'A/R' in l or 'Account No' in l:
                    part = l.split('A/R')[0].strip()
                    if part: name_parts.append(part)
                    continue
                if 'Print Date' in l or 'Page No' in l: break
                if len(l) < 40 and not any(kw in l for kw in ['Date', 'Folio', 'Debit', 'Credit']):
                    name_parts.append(l)
            return ' '.join(name_parts) if name_parts else ''
    except:
        return ''

_SOA_HEADER_COLS = ['Date', 'Folio', 'Description', 'Arrival', 'Departure', 'Voucher', 'Debit', 'Credit', 'Balance']

def _detect_soa_col_bounds(pdf_bytes):
    """Detect column boundaries dynamically from the header row AND actual data rows."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:2]:
                words = page.extract_words()
                rows_by_y = defaultdict(list)
                for w in words:
                    rows_by_y[round(w['top'] / 3) * 3].append(w)

                # ── Step 1: find header row to get approximate column x0 positions ──
                col = None
                header_y = None
                for y in sorted(rows_by_y.keys()):
                    row_words = rows_by_y[y]
                    texts = [w['text'] for w in row_words]
                    if 'Departure' in texts and 'Voucher' in texts:
                        c = {w['text']: w['x0'] for w in row_words if w['text'] in _SOA_HEADER_COLS}
                        if all(k in c for k in ['Date','Folio','Description','Arrival','Departure','Voucher','Debit']):
                            col = c
                            header_y = y
                            break

                if not col:
                    continue

                def mid(a, b): return (col[a] + col[b]) / 2

                # ── Step 2: scan data rows to find actual Arrival date x0 positions ──
                # Look for dd/mm/yy values between Description and Departure columns
                arr_data_x0s = []
                for y in sorted(rows_by_y.keys()):
                    if y <= header_y or y > 715:
                        continue
                    for w in rows_by_y[y]:
                        if (re.match(r'^\d{2}/\d{2}/\d{2}$', w['text']) and
                                col['Description'] < w['x0'] < col['Departure']):
                            arr_data_x0s.append(w['x0'])
                    if len(arr_data_x0s) >= 10:
                        break  # enough samples

                # DESC_MAX = 2px before the leftmost actual arrival date found
                # This is truly data-driven — adapts to each PDF automatically
                if arr_data_x0s:
                    desc_max = min(arr_data_x0s) - 2
                else:
                    desc_max = col['Arrival'] - 5  # fallback if no date data found

                return dict(
                    DATE_MAX      = mid('Date', 'Folio'),
                    FOLIO_MAX     = mid('Folio', 'Description'),
                    DESC_MAX      = desc_max,
                    ARR_MAX       = mid('Arrival', 'Departure'),
                    DEP_MAX       = mid('Departure', 'Voucher'),
                    VCH_MAX       = mid('Voucher', 'Debit'),
                    DEBIT_X1_MAX  = col.get('Credit', col['Debit'] + 50),
                    CREDIT_X1_MAX = col.get('Balance', col.get('Credit', col['Debit'] + 80) + 50),
                )
    except:
        pass
    # Fallback (Batch-style layout)
    return dict(DATE_MAX=74, FOLIO_MAX=133, DESC_MAX=218, ARR_MAX=298,
                DEP_MAX=356, VCH_MAX=411, DEBIT_X1_MAX=490, CREDIT_X1_MAX=546)

def extract_statement_rows(pdf_bytes):
    b = _detect_soa_col_bounds(pdf_bytes)
    DATE_MAX      = b['DATE_MAX']
    FOLIO_MAX     = b['FOLIO_MAX']
    DESC_MAX      = b['DESC_MAX']
    ARR_MAX       = b['ARR_MAX']
    DEP_MAX       = b['DEP_MAX']
    VCH_MAX       = b['VCH_MAX']
    DEBIT_X1_MAX  = b['DEBIT_X1_MAX']
    CREDIT_X1_MAX = b['CREDIT_X1_MAX']

    HEADER_WORDS = {'description', 'voucher', 'folio', 'arrival', 'departure',
                    'debit', 'credit', 'balance', 'date'}

    all_rows = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            rows_by_y = defaultdict(list)
            for w in words:
                rows_by_y[round(w['top'] / 3) * 3].append(w)

            # ── Pass 1: identify primary rows (date + folio) ──
            primary_ys   = []
            primary_data = {}

            # Pre-scan footer boundary — stop before Aging/Balance-Due/Bank sections
            footer_y = 715
            for y in sorted(rows_by_y.keys()):
                if y < 195:
                    continue
                rw_texts = [w['text'] for w in rows_by_y[y]]
                is_footer = (
                    # "Balance Due" on same line (Keytel format)
                    ('Balance' in rw_texts and 'Due' in rw_texts) or
                    # THB currency marker (appears on Balance Due amount line)
                    '(THB)' in rw_texts or
                    # Aging Summary section header
                    'Aging' in rw_texts or
                    # Bank Details section header
                    ('Bank' in rw_texts and 'Details' in rw_texts) or
                    # Standalone "Balance" split across lines (Luxury Escapes format)
                    ('Balance' in rw_texts and len(rw_texts) <= 2 and
                     not any(re.match(r'\d{2}/\d{2}/\d{2}', t) for t in rw_texts))
                )
                if is_footer:
                    footer_y = y - 1
                    break

            for y in sorted(rows_by_y.keys()):
                if y < 195 or y > footer_y:
                    continue
                ws_words = sorted(rows_by_y[y], key=lambda w: w['x0'])
                date_w, folio_w, desc_w = [], [], []
                arr_w, dep_w, vch_w = [], [], []
                debit_w, credit_w, bal_w = [], [], []

                for w in ws_words:
                    x0, x1, t = w['x0'], w['x1'], w['text']
                    if x0 < DATE_MAX:          date_w.append(t)
                    elif x0 < FOLIO_MAX:       folio_w.append(t)
                    elif x0 < DESC_MAX:        desc_w.append(t)
                    elif x0 < ARR_MAX:         arr_w.append(t)
                    elif x0 < DEP_MAX:         dep_w.append(t)
                    elif x0 < VCH_MAX:         vch_w.append(t)
                    else:
                        if x1 <= DEBIT_X1_MAX:    debit_w.append(t)
                        elif x1 <= CREDIT_X1_MAX: credit_w.append(t)
                        else:                     bal_w.append(t)

                dv = ' '.join(date_w)
                fv = ' '.join(folio_w)
                if dv and fv and re.match(r'\d{2}/\d{2}/\d{2}', dv):
                    primary_ys.append(y)
                    primary_data[y] = {
                        'date': dv, 'folio': fv,
                        'desc': ' '.join(desc_w), 'arrival': ' '.join(arr_w),
                        'departure': ' '.join(dep_w), 'voucher': ' '.join(vch_w),
                        'debit': ' '.join(debit_w), 'credit': ' '.join(credit_w),
                        'balance': ' '.join(bal_w),
                    }

            if not primary_ys:
                continue

            # ── Pass 2: collect continuations with y positions ──
            # Use defaultdict keyed by nearest_py; each field stores [(y, text), ...]
            cont = defaultdict(lambda: {'desc': [], 'voucher': [],
                                        'debit': [], 'credit': [], 'balance': []})

            for y in sorted(rows_by_y.keys()):
                if y < 195 or y > footer_y:
                    continue
                if y in primary_data:
                    continue  # already handled as a primary row

                # Find nearest primary using midpoint rule
                nearest_py = None
                if y < primary_ys[0]:
                    # Before first primary — belongs to the first primary
                    nearest_py = primary_ys[0]
                else:
                    for i in range(len(primary_ys) - 1):
                        mid = (primary_ys[i] + primary_ys[i + 1]) / 2
                        if y <= mid:
                            nearest_py = primary_ys[i]
                            break
                    if nearest_py is None:
                        nearest_py = primary_ys[-1]

                # Skip if too far from nearest primary (aging/summary section)
                if abs(y - nearest_py) > 60:
                    continue

                ws_words = sorted(rows_by_y[y], key=lambda w: w['x0'])
                desc_w, vch_w, debit_w, credit_w, bal_w = [], [], [], [], []
                for w in ws_words:
                    x0, x1, t = w['x0'], w['x1'], w['text']
                    if FOLIO_MAX <= x0 < DESC_MAX:
                        desc_w.append(t)
                    elif DEP_MAX <= x0 < VCH_MAX:
                        vch_w.append(t)
                    elif x0 >= VCH_MAX:
                        if x1 <= DEBIT_X1_MAX:    debit_w.append(t)
                        elif x1 <= CREDIT_X1_MAX: credit_w.append(t)
                        else:                     bal_w.append(t)

                desc_cont   = ' '.join(desc_w).strip()
                vch_cont    = ' '.join(vch_w).strip()
                debit_cont  = ''.join(debit_w)
                credit_cont = ''.join(credit_w)
                bal_cont    = ''.join(bal_w)

                # Skip repeated header rows
                if desc_cont.lower() in HEADER_WORDS or vch_cont.lower() in HEADER_WORDS:
                    continue

                if desc_cont:   cont[nearest_py]['desc'].append((y, desc_cont))
                if vch_cont:    cont[nearest_py]['voucher'].append((y, vch_cont))
                if debit_cont:  cont[nearest_py]['debit'].append((y, debit_cont))
                if credit_cont: cont[nearest_py]['credit'].append((y, credit_cont))
                if bal_cont:    cont[nearest_py]['balance'].append((y, bal_cont))

            # ── Build final field values in correct y order ──
            for py in primary_ys:
                row = primary_data[py]

                # Voucher: sort all fragments (primary + continuations) by y,
                # then concatenate WITHOUT spaces — long references wrap mid-token
                vch_all = sorted(cont[py]['voucher'] + [(py, row['voucher'])],
                                 key=lambda x: x[0])
                row['voucher'] = ''.join(v for _, v in vch_all if v)

                # Desc: sort by y, join WITH spaces (name/text wraps with natural spaces)
                desc_all = sorted(cont[py]['desc'] + [(py, row['desc'])],
                                  key=lambda x: x[0])
                row['desc'] = ' '.join(d for _, d in desc_all if d)

                # Debit/Credit/Balance: sort by y, concatenate WITHOUT spaces
                # (digits split across lines, e.g. "151,130.0" + "0" → "151,130.00")
                for field in ('debit', 'credit', 'balance'):
                    parts = sorted(cont[py][field] + [(py, row[field])],
                                   key=lambda x: x[0])
                    row[field] = ''.join(v for _, v in parts if v)

                all_rows.append(row)

    return all_rows

def convert_statement(pdf_bytes, client_name, print_date):
    rows = extract_statement_rows(pdf_bytes)

    if not rows:
        raise ValueError("ไม่พบข้อมูลใน PDF — กรุณาตรวจสอบว่าเป็น Statement of Account ที่รองรับ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement of Account"

    white_f  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    total_f  = Font(name='Arial', bold=True, size=10)
    title_f  = Font(name='Arial', bold=True, size=13, color="1F4E79")
    sub_f    = Font(name='Arial', bold=True, size=11, color="2E75B6")
    c_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    r_align  = Alignment(horizontal='right',  vertical='center')
    thin     = Side(style='thin',   color='BFBFBF')
    medium   = Side(style='medium', color='2E75B6')
    t_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    tot_fill = PatternFill("solid", fgColor="D6E4F0")
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    num_fmt  = '#,##0.00;[Red](#,##0.00);"-"'

    ws.merge_cells('A1:I1')
    ws['A1'] = "STATEMENT OF ACCOUNT"
    ws['A1'].font = title_f; ws['A1'].alignment = c_align
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:G2')
    ws['A2'] = client_name
    ws['A2'].font = sub_f; ws['A2'].alignment = l_align
    ws['H2'] = "Print Date:"
    ws['H2'].font = Font(name='Arial', bold=True, size=9); ws['H2'].alignment = r_align
    ws['I2'] = print_date
    ws['I2'].font = Font(name='Arial', size=9); ws['I2'].alignment = c_align
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 5

    headers = ['Date', 'Folio', 'Description', 'Arrival', 'Departure', 'Voucher', 'Debit', 'Credit', 'Balance']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = white_f; c.fill = hdr_fill
        c.alignment = c_align; c.border = t_border
    ws.row_dimensions[4].height = 22

    for i, row in enumerate(rows):
        er = 5 + i
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            row['date'], row['folio'], row['desc'],
            row['arrival'], row['departure'], row['voucher'],
            clean_num(row['debit']), clean_num(row['credit']), clean_num(row['balance'])
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=er, column=col, value=val)
            c.font = Font(name='Arial', size=9)
            c.fill = fill; c.border = t_border
            if col == 3:        c.alignment = l_align
            elif col >= 7:
                c.alignment = r_align
                if val is not None: c.number_format = num_fmt
            else:               c.alignment = c_align
        ws.row_dimensions[er].height = 16

    tr = 5 + len(rows)
    td = sum(clean_num(r['debit'])   or 0 for r in rows)
    tc = sum(clean_num(r['credit'])  or 0 for r in rows)
    tb = sum(clean_num(r['balance']) or 0 for r in rows)

    ws.merge_cells(f'A{tr}:F{tr}')
    c = ws.cell(row=tr, column=1, value='รวมทั้งสิ้น / Grand Total')
    c.font = total_f; c.alignment = c_align
    c.fill = tot_fill; c.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    for col in range(2, 7):
        c = ws.cell(row=tr, column=col)
        c.fill = tot_fill; c.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    for col, val in [(7, td), (8, tc if tc else None), (9, tb)]:
        c = ws.cell(row=tr, column=col, value=val)
        c.fill = tot_fill; c.font = total_f
        c.border = Border(left=thin, right=thin, top=medium, bottom=medium)
        c.alignment = r_align
        if val is not None: c.number_format = num_fmt
    ws.row_dimensions[tr].height = 22

    col_widths = [12, 9, 42, 12, 12, 22, 16, 14, 16]
    for i, cw in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = cw

    ws.freeze_panes = 'A5'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(rows), td, tc, tb

def get_print_date(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
            m = re.search(r'Print\s+Date\s+(\d{2}/\d{2}/\d{2})', text)
            if m: return m.group(1)
    except:
        pass
    return ''


# ═══ AR Aging support (Aging Summary + AR Detailed Aging) ═══════════
# Handles normal searchable PDFs via pdfplumber words/chars, and
# "garbled" Type3-bitmap-font PDFs via a deterministic raster reader
# (template matching against an Arial glyph bank in aging_glyph_bank.npz).
import os

try:
    import numpy as _np
    import pypdfium2 as _pdfium
    from PIL import Image as _PILImage, ImageDraw as _PILDraw, ImageFont as _PILFont
    AGING_RASTER_OK = True
except Exception:
    AGING_RASTER_OK = False

_AG_CHARS = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789.,/:%()-&'"
_AG_N = 24
_AG_SHEAR = 0.2125
_AG_LIB_FONTS = [
    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    '/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf',
    '/usr/share/fonts/truetype/liberation/LiberationSans-BoldItalic.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
]

def _ag_norm(a):
    ys, xs = _np.nonzero(a)
    if len(xs) == 0: return None, 1.0
    a = a[ys.min():ys.max()+1, xs.min():xs.max()+1]
    h, w = a.shape
    side = max(h, w)
    sq = _np.zeros((side, side), dtype=_np.float32)
    sq[(side-h)//2:(side-h)//2+h, (side-w)//2:(side-w)//2+w] = a
    im = _PILImage.fromarray((sq*255).astype(_np.uint8)).resize((_AG_N, _AG_N), _PILImage.LANCZOS)
    return _np.asarray(im, dtype=_np.float32)/255.0, w/h

_AG_TPL = None
def _ag_templates():
    global _AG_TPL
    if _AG_TPL is not None: return _AG_TPL
    labels, arrs, aspects = [], [], []
    covered = set()
    here = os.path.dirname(os.path.abspath(__file__))
    for p in (os.path.join(here, 'aging_glyph_bank.npz'), 'aging_glyph_bank.npz'):
        if os.path.exists(p):
            z = _np.load(p, allow_pickle=False)
            for ch, arr, ar in zip(z['chars'], z['arrs'], z['aspects']):
                labels.append(str(ch)); arrs.append(arr); aspects.append(float(ar))
                if len(str(ch)) == 1: covered.add(str(ch))
            break
    for fp in _AG_LIB_FONTS:
        try: font = _PILFont.truetype(fp, 96)
        except Exception: continue
        italic = 'Italic' in fp
        for ch in _AG_CHARS:
            if ch in covered and not italic: continue
            img = _PILImage.new('L', (220, 200), 0)
            _PILDraw.Draw(img).text((60, 40), ch, fill=255, font=font)
            t, ar = _ag_norm((_np.asarray(img) > 128).astype(_np.float32))
            if t is not None:
                labels.append(ch); arrs.append(t); aspects.append(ar)
    _AG_TPL = (_np.array(labels), _np.stack(arrs).astype(_np.float32),
               _np.array(aspects, dtype=_np.float32))
    return _AG_TPL

def _ag_classify(g):
    t, ar = _ag_norm(g)
    if t is None: return [(' ', 0.0)]
    labels, T, A = _ag_templates()
    d = _np.abs(T - t[None]).mean(axis=(1, 2)) + 0.35*_np.abs(A - ar)/_np.maximum(_np.maximum(A, ar), 1e-6)
    order = _np.argsort(d)
    out, seen = [], set()
    for i in order:
        ch = labels[i]
        if ch in seen: continue
        seen.add(ch); out.append((ch, float(-d[i])))
        if len(out) >= 10: break
    return out

def _ag_spans(mask):
    out, s = [], None
    for i, v in enumerate(mask):
        if v and s is None: s = i
        elif not v and s is not None: out.append((s, i)); s = None
    if s is not None: out.append((s, len(mask)))
    return out

def _ag_split_wide(band, a, b, med_w, depth=0):
    w = b - a
    if med_w <= 0 or w < 1.45*med_w or depth > 3:
        return [(a, b)]
    whole = _ag_classify(band[:, a:b])[0][1]
    prof = band[:, a:b].sum(axis=0)
    lo = a + max(int(0.25*med_w), 2)
    hi = b - max(int(0.25*med_w), 2)
    if hi <= lo: return [(a, b)]
    cands = list(range(lo, hi, 2))
    thresh = _np.percentile(prof, 45)
    cands = [c for c in cands if prof[c-a] <= thresh] or cands
    best_cut, best_s = None, -9.0
    for c in cands:
        s2 = min(_ag_classify(band[:, a:c])[0][1], _ag_classify(band[:, c:b])[0][1])
        if s2 > best_s: best_s, best_cut = s2, c
    if best_cut is None or best_s <= whole + 0.015:
        return [(a, b)]
    return (_ag_split_wide(band, a, best_cut, med_w, depth+1) +
            _ag_split_wide(band, best_cut, b, med_w, depth+1))

def _ag_fix_bars(txt, wg):
    chars = list(txt)
    letters = [c for c in chars if len(c) == 1 and c.isalpha()]
    n_upper = sum(1 for c in letters if c.isupper())
    n_lower = len(letters) - n_upper
    for i, (c, g) in enumerate(zip(chars, wg)):
        if c not in 'Ili': continue
        if g.get('ncomp', 1) >= 2:
            chars[i] = 'i'
        elif i == 0 and len(chars) > 1 and chars[1].islower():
            chars[i] = 'I'
        elif n_upper > n_lower:
            chars[i] = 'I'
        else:
            chars[i] = 'l'
    return ''.join(chars)

def _ag_is_garbled(text):
    if not text: return True
    return text.count('(cid:') > 10

def _ag_rotation(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            chars = pdf.pages[0].chars
            if not chars: return False
            return sum(1 for c in chars if not c.get('upright', True)) > len(chars)/2
    except Exception:
        return False

def raster_words_pages(pdf_bytes, rotate, scale=10, progress_cb=None):
    doc = _pdfium.PdfDocument(io.BytesIO(pdf_bytes))
    pages_out = []
    for pi in range(len(doc)):
        if progress_cb:
            try: progress_cb(pi + 1, len(doc))
            except Exception: pass
        pil = doc[pi].render(scale=scale).to_pil()
        if rotate: pil = pil.transpose(_PILImage.ROTATE_270)
        arr = (_np.asarray(pil.convert('L')) < 160).astype(_np.float32)
        Himg, Wimg = arr.shape
        rowsum = arr.sum(axis=1)
        for ri in _np.nonzero(rowsum > 0.25*Wimg)[0]:
            for (a, b) in _ag_spans(arr[ri] > 0):
                if b - a > 15*scale:
                    arr[ri, a:b] = 0.0
        bands0 = _ag_spans(arr.sum(axis=1) > 0)
        bands = []
        for b in bands0:
            if bands and b[0]-bands[-1][1] <= 2 and min(b[1]-b[0], bands[-1][1]-bands[-1][0]) <= 3:
                bands[-1] = (bands[-1][0], b[1])
            else: bands.append(list(b))
        max_line_h = 11.0*scale
        bands2 = []
        for (y0, y1) in bands:
            if y1-y0 > max_line_h:
                sub = arr[y0:y1]
                rs = sub.sum(axis=1)
                weak = rs <= max(2, 0.0015*Wimg)
                inner = _ag_spans(~weak)
                if len(inner) > 1:
                    bands2.extend((y0+a, y0+b) for a, b in inner)
                    continue
                mid0, mid1 = int(0.25*(y1-y0)), int(0.75*(y1-y0))
                cut = mid0 + int(_np.argmin(rs[mid0:mid1]))
                nz = rs[rs > 0]
                if len(nz) and rs[cut] < 0.25*_np.median(nz):
                    bands2.append((y0, y0+cut)); bands2.append((y0+cut, y1))
                    continue
            bands2.append((y0, y1))
        words = []
        for (y0, y1) in bands2:
            band = arr[y0:y1, :]
            spans = _ag_spans(band.sum(axis=0) > 0)
            if not spans: continue
            widths = [b-a for a, b in spans]
            big = [w for w in widths if w > 2]
            med_w = float(_np.median(big)) if big else float(_np.median(widths))
            spans = [p for s in spans for p in _ag_split_wide(band, s[0], s[1], med_w)]

            def _build(bnd, spans_):
                out = []
                for (a, b) in spans_:
                    g = bnd[:, a:b]
                    ys, xs = _np.nonzero(g)
                    if len(xs) == 0: continue
                    gtop, gbot = int(ys.min()), int(ys.max())
                    gh = gbot-gtop+1
                    if gh <= 3 and (b-a) > 12*gh and (b-a) > 150:
                        continue
                    ncomp = len(_ag_spans(g.sum(axis=1) > 0))
                    out.append({'a': a, 'b': b, 'top': gtop, 'bot': gbot, 'h': gh,
                                'ncomp': ncomp, 'cands': _ag_classify(g)})
                return out
            glyphs = _build(band, spans)
            if not glyphs: continue
            mean_s = float(_np.mean([g['cands'][0][1] for g in glyphs]))
            if mean_s < -0.20:
                h = band.shape[0]
                sheared = _np.zeros_like(band)
                for r in range(h):
                    off = int(round(_AG_SHEAR*(h-1-r)))
                    if off > 0: sheared[r, :-off or None] = band[r, off:]
                    else: sheared[r] = band[r]
                spans_i = _ag_spans(sheared.sum(axis=0) > 0)
                if spans_i:
                    w_i = [b-a for a, b in spans_i]
                    big_i = [w for w in w_i if w > 2]
                    med_i = float(_np.median(big_i)) if big_i else float(_np.median(w_i))
                    spans_i = [p for sp in spans_i for p in _ag_split_wide(sheared, sp[0], sp[1], med_i)]
                    glyphs_i = _build(sheared, spans_i)
                    if glyphs_i:
                        mean_i = float(_np.mean([g['cands'][0][1] for g in glyphs_i]))
                        if mean_i > mean_s + 0.03:
                            glyphs, band = glyphs_i, sheared
            alnum_hs = [g['h'] for g in glyphs
                        if g['cands'][0][0][:1].isalnum()]
            cap_h = float(_np.percentile(alnum_hs, 75)) if alnum_hs \
                    else max(g['h'] for g in glyphs)
            bots = [g['bot'] for g in glyphs if g['h'] > 0.5*cap_h]
            baseline = float(_np.median(bots)) if bots else max(g['bot'] for g in glyphs)
            for g in glyphs:
                ch = g['cands'][0][0]
                if len(ch) == 1 and ch in ".,'":
                    clipped = g['bot'] >= (y1 - y0) - 2
                    if g['top'] < baseline - 0.55*cap_h: ch = "'"
                    elif clipped: ch = g['cands'][0][0]
                    elif g['bot'] > baseline + max(2, 0.06*cap_h): ch = ','
                    else: ch = '.'
                elif len(ch) == 1 and ch.lower() in 'cosuvwxz':
                    ch = ch.upper() if g['h'] >= 0.85*cap_h else ch.lower()
                g['ch'] = ch
                g['cap_h'] = cap_h; g['baseline'] = baseline
            gap_thr = max(3.0, 0.5*med_w)
            wlist, cur = [], [glyphs[0]]
            for g in glyphs[1:]:
                if g['a'] - cur[-1]['b'] > gap_thr: wlist.append(cur); cur = [g]
                else: cur.append(g)
            wlist.append(cur)
            for wg in wlist:
                txt = _ag_fix_bars(''.join(g['ch'] for g in wg), wg)
                words.append({'text': txt,
                              'x0': wg[0]['a']/scale, 'x1': wg[-1]['b']/scale,
                              'top': y0/scale, 'bottom': y1/scale, 'glyphs': wg})
        pages_out.append(words)
    return pages_out

def _ag_restrict(word, allowed):
    if not word.get('glyphs'):
        return word['text']
    out = []
    for g in word['glyphs']:
        best = None
        for ch, s in g['cands']:
            if len(ch) != 1: continue
            c = ch
            if c in ".,'":
                c = ',' if g['bot'] > g['baseline'] + max(2, 0.06*g['cap_h']) else '.'
            if c in allowed: best = c; break
        if best is None:
            for ch, s in g['cands']:
                if len(ch) != 1: continue
                if ch.upper() in allowed: best = ch.upper(); break
                if ch.lower() in allowed: best = ch.lower(); break
        out.append(best or '?')
    return ''.join(out)

# ─── shared parsing helpers ──────────────────────────────────────────
def _ag_lines(words, tol=3.5):
    if not words: return []
    ws = sorted(words, key=lambda w: (w['top']+w['bottom'])/2)
    lines, cur = [], [ws[0]]
    for w in ws[1:]:
        c0 = (cur[-1]['top']+cur[-1]['bottom'])/2
        c1 = (w['top']+w['bottom'])/2
        if c1 - c0 <= tol: cur.append(w)
        else: lines.append(sorted(cur, key=lambda x: x['x0'])); cur = [w]
    lines.append(sorted(cur, key=lambda x: x['x0']))
    return lines

def _ag_glue(line):
    """normalized keyword string: uppercase, no spaces, numeric tokens dropped"""
    toks = [w['text'] for w in line
            if not re.match(r'^-?[\d.,:%/]+$', w['text'])]
    s = ''.join(toks).upper().replace('1', 'I').replace('0', 'O')
    return re.sub(r'[^A-Z/&]', '', s)

def _ag_amount(s):
    """parse an amount that always carries 2 decimals; immune to ,/. confusion"""
    s = s.replace(' ', '').replace('%', '')
    neg = s.startswith('-') or s.endswith('-')
    digits = re.sub(r'\D', '', s)
    if not digits: return None
    v = int(digits) / 100.0
    return -v if neg else v

def _ag_merge_line_numbers(line):
    """merge '-' into following numeric word; merge adjacent numeric fragments"""
    def numish(t):
        return bool(re.match(r'^[\d.,%]+$', t)) or t == '-'
    out = []
    i = 0
    while i < len(line):
        w = dict(line[i])
        if w['text'] == '-' and i+1 < len(line) and \
           re.match(r'^[\d.,]', line[i+1]['text']) and line[i+1]['x0'] - w['x1'] < 8:
            nxt = line[i+1]
            w = {'text': '-' + nxt['text'], 'x0': w['x0'], 'x1': nxt['x1'],
                 'top': w['top'], 'bottom': w['bottom']}
            i += 2
        else:
            i += 1
        if out and numish(out[-1]['text']) and numish(w['text']) and \
           w['x0'] - out[-1]['x1'] < 3.4 and not w['text'].startswith('-'):
            out[-1] = {'text': out[-1]['text'] + w['text'],
                       'x0': out[-1]['x0'], 'x1': w['x1'],
                       'top': out[-1]['top'], 'bottom': out[-1]['bottom']}
        else:
            out.append(w)
    return out

def _ag_fix_accno(s):
    s = re.sub(r'[^A-Za-z0-9]', '', s).upper()
    if len(s) < 5: return s
    head, tail = s[:-4], s[-4:]
    dmap = {'O':'0','Q':'0','I':'1','L':'1','S':'5','B':'8','Z':'2','G':'6','D':'0'}
    lmap = {'0':'O','1':'I','5':'S','8':'B','2':'Z'}
    tail = ''.join(dmap.get(c, c) for c in tail)
    head = ''.join(lmap.get(c, c) for c in head)
    if re.match(r'^[A-Z]{2,6}$', head) and re.match(r'^\d{4}$', tail):
        return head + tail
    return s

_AG_BUCKETS = ['Up to 30', '31 - 60', '61 - 90', '91 - 120', '121 - 150',
               '151 and Over', 'Total']

def _ag_find_header(lines, need_words, anchor_last_words):
    """find header line containing need_words; return (idx, anchors[x1 list], meta_x)"""
    for idx, line in enumerate(lines):
        texts = [w['text'] for w in line]
        if all(any(t == nw for t in texts) for nw in need_words):
            anchors = []
            for aw in anchor_last_words:
                cand = [w for w in line if w['text'] == aw]
                if not cand: return None
                anchors.append(cand[-1]['x1'])
            return idx, anchors, line
    return None

# ─── Aging Summary parser ────────────────────────────────────────────
def _ag_get_words_pages(pdf_bytes):
    """words per page: pdfplumber for clean PDFs, raster engine for garbled"""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text0 = pdf.pages[0].extract_text() or ''
        if not _ag_is_garbled(text0):
            return [p.extract_words() for p in pdf.pages], 'text'
    if not AGING_RASTER_OK:
        raise ValueError("PDF นี้ฝังฟอนต์แบบพิเศษ (Type3) ต้องใช้ตัวอ่าน raster — "
                         "กรุณาติดตั้ง pypdfium2, numpy, Pillow")
    rot = _ag_rotation(pdf_bytes)
    return raster_words_pages_cached(pdf_bytes, rot), 'raster'

def parse_aging_summary(pdf_bytes):
    pages, source = _ag_get_words_pages(pdf_bytes)
    meta = {'report': 'Aging Summary for All Types', 'source': source,
            'property': '', 'print_date': '', 'print_time': '',
            'business_date': '', 'age_credits': ''}
    items = []
    for pi, words in enumerate(pages):
        lines = _ag_lines(words)
        hdr = _ag_find_header(lines, ['Credit', 'Limit'],
                              ['Limit', '30', '60', '90', '120', '150', 'Over', 'Total'])
        if hdr is None: continue
        hdr_i, anchors, hline = hdr
        climit_a, bucket_a = anchors[0], anchors[1:]
        acc_words = [w for w in hline if w['text'] == 'Account']
        accno_x0 = acc_words[-1]['x0'] if len(acc_words) >= 2 else hline[0]['x0'] + 115
        name_end = accno_x0 - 2
        num_zone = accno_x0 + 48
        if pi == 0:
            # meta from lines above header
            for line in lines[:hdr_i]:
                raw = ' '.join(w['text'] for w in line)
                if re.match(r'^\d{2}:\d{2}$', raw.strip()):
                    meta['print_time'] = raw.strip(); continue
                glue = _ag_glue(line)
                if 'AGECREDIT' in glue:
                    meta['age_credits'] = line[-1]['text']; continue
                if 'SUMMARY' in glue or 'AGING' in glue or 'TYPES' in glue:
                    continue
                letters = sum(1 for ch in raw if ch.isalpha())
                if letters >= 8 and not meta['property']:
                    toks = [w['text'] for w in line
                            if re.search(r'[A-Za-z0-9]', w['text'])]
                    if toks and re.match(r'^\d{2}/\d{2}/\d{2}$', toks[-1]):
                        meta['print_date'] = toks[-1]; toks = toks[:-1]
                    meta['property'] = ' '.join(toks)
        footer = False
        for li, line in enumerate(lines):
            if li <= hdr_i: continue
            glue = _ag_glue(line)
            raw = ' '.join(w['text'] for w in line)
            if footer or 'FILTER' in glue or 'BUSINESSDATE' in glue or ('PAGE' in glue and 'OF' in glue):
                footer = True
                m = re.search(r'(\d{2}/\d{2}/\d{2})', raw)
                if 'BUSINESSDATE' in glue and m and not meta['business_date']:
                    meta['business_date'] = m.group(1)
                continue
            if glue in ('/LOC', 'LOC', ''):
                continue
            # numbers
            merged = _ag_merge_line_numbers(line)
            frag = defaultdict(list)   # anchor idx (-1=climit, 0..6 buckets) -> texts
            plain = []
            for w in merged:
                t = w['text']
                isnum = bool(re.match(r'^-?[\d.,]+%?$', t)) or t == '%'
                if isnum and (w['x0'] > name_end):
                    all_a = [climit_a] + bucket_a
                    ds = [abs(w['x1'] - a) for a in all_a]
                    bi = int(_np.argmin(ds)) if AGING_RASTER_OK else ds.index(min(ds))
                    if ds[bi] <= 28:
                        frag[bi-1].append((w['x0'], t)); continue
                plain.append(w)
            vals = {}
            for k, lst in frag.items():
                vals[k] = ''.join(t for _, t in sorted(lst))
            bucket_vals = [_ag_amount(vals[k]) if k in vals else None for k in range(7)]
            has_nums = any(v is not None for v in bucket_vals)
            name_w = [w for w in plain if w['x1'] <= name_end + 6]
            accno_w = [w for w in plain if w['x0'] > name_end and w['x0'] < num_zone
                       and re.match(r'^[A-Za-z0-9]+$', w['text'])]
            name_txt = ' '.join(w['text'] for w in name_w).strip()
            # classify
            if re.match(r'^TOTAL[FT]ORACCOUNTTYPE', glue):
                toks = [w['text'] for w in plain]
                code = toks[-1] if toks else ''
                items.append({'kind': 'type_total', 'code': code, 'vals': bucket_vals})
            elif re.match(r'^NOO[FT]ACCOUNT', glue):
                cnt = None
                if -1 in vals:
                    d = re.sub(r'\D', '', vals[-1])
                    cnt = int(d) if d else None
                items.append({'kind': 'pct', 'count': cnt, 'vals': bucket_vals})
            elif re.match(r'^TOTALA/?RLEDGER', glue):
                cnt = None
                if -1 in vals:
                    d = re.sub(r'\D', '', vals[-1])
                    cnt = int(d) if d else None
                items.append({'kind': 'ledger_total', 'count': cnt, 'vals': bucket_vals})
            elif re.match(r'^TOTALACCR[UO]A[LI]S', glue):
                items.append({'kind': 'accruals_total', 'vals': bucket_vals})
            elif glue.startswith('GRANDTOTAL'):
                items.append({'kind': 'grand_total', 'vals': bucket_vals})
            elif glue.startswith('ACCOUNTTYPE'):
                toks = [w['text'] for w in line]
                low = [t.upper() for t in toks]
                rest = toks[2:] if low[:2] == ['ACCOUNT', 'TYPE'] else toks[1:]
                code = rest[0] if rest else ''
                desc = ' '.join(rest[1:])
                items.append({'kind': 'type_header', 'code': code, 'desc': desc})
            elif re.match(r'^ACCR[UO]A[LI]S$', glue) and not has_nums:
                items.append({'kind': 'section', 'label': 'Accruals'})
            elif re.match(r'^A/?RLEDGER$', glue) and not has_nums:
                items.append({'kind': 'section', 'label': 'A/R Ledger'})
            elif has_nums:
                accno = ''
                if accno_w:
                    joined = ''.join(w['text'] for w in sorted(accno_w, key=lambda x: x['x0']))
                    if len(accno_w) == 1 and accno_w[0].get('glyphs'):
                        joined = _ag_restrict(accno_w[0], set('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'))
                    accno = _ag_fix_accno(joined)
                climit = _ag_amount(vals[-1]) if -1 in vals else None
                items.append({'kind': 'account', 'name': name_txt, 'accno': accno,
                              'climit': climit, 'vals': bucket_vals})
            elif name_txt and items and items[-1]['kind'] == 'account':
                items[-1]['name'] = (items[-1]['name'] + ' ' + name_txt).strip()
    if not any(it['kind'] == 'account' for it in items):
        raise ValueError("ไม่พบข้อมูลบัญชีใน PDF — กรุณาตรวจสอบว่าเป็น Aging Summary ที่รองรับ")
    if meta.get('source') == 'raster':
        _ag_title_refine(pdf_bytes, meta)
    return meta, items

def validate_aging_summary(items):
    issues, checks = [], 0
    open_accounts = []
    running_type_totals = []
    accrual_flag = False
    ledger_vals = None
    accrual_vals = None
    pre = []   # type totals before Accruals section
    post = []
    for it in items:
        if it['kind'] == 'account':
            checks += 1
            b = it['vals']
            if None not in b[:7]:
                if abs(sum(x for x in b[:6]) - b[6]) > 0.015:
                    issues.append(f"แถว {it['name']}: ผลรวม bucket ≠ Total ({sum(b[:6]):,.2f} vs {b[6]:,.2f})")
            open_accounts.append(b)
        elif it['kind'] == 'type_total':
            checks += 1
            for ci in range(7):
                s = sum((a[ci] or 0) for a in open_accounts)
                v = it['vals'][ci] or 0
                if abs(s - v) > 0.015:
                    issues.append(f"Total type {it['code']} คอลัมน์ {ci+1}: {s:,.2f} ≠ {v:,.2f}")
            (post if accrual_flag else pre).append(it['vals'])
            open_accounts = []
        elif it['kind'] == 'section' and it['label'] == 'Accruals':
            accrual_flag = True
        elif it['kind'] == 'ledger_total':
            checks += 1
            ledger_vals = it['vals']
            for ci in range(7):
                s = sum((t[ci] or 0) for t in pre)
                if abs(s - (it['vals'][ci] or 0)) > 0.015:
                    issues.append(f"Total AR Ledger คอลัมน์ {ci+1}: {s:,.2f} ≠ {(it['vals'][ci] or 0):,.2f}")
        elif it['kind'] == 'accruals_total':
            checks += 1
            accrual_vals = it['vals']
            for ci in range(7):
                s = sum((t[ci] or 0) for t in post)
                if abs(s - (it['vals'][ci] or 0)) > 0.015:
                    issues.append(f"Total Accruals คอลัมน์ {ci+1}: {s:,.2f} ≠ {(it['vals'][ci] or 0):,.2f}")
        elif it['kind'] == 'grand_total':
            checks += 1
            if ledger_vals and accrual_vals:
                for ci in range(7):
                    s = (ledger_vals[ci] or 0) + (accrual_vals[ci] or 0)
                    if abs(s - (it['vals'][ci] or 0)) > 0.015:
                        issues.append(f"Grand Total คอลัมน์ {ci+1}: {s:,.2f} ≠ {(it['vals'][ci] or 0):,.2f}")
    return checks, issues

# ─── AR Detailed Aging parser ────────────────────────────────────────
def _ag_get_chars_pages(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text0 = pdf.pages[0].extract_text() or ''
        if not _ag_is_garbled(text0):
            out = []
            for p in pdf.pages:
                out.append([{'text': c['text'], 'x0': c['x0'], 'x1': c['x1'],
                             'top': c['top'], 'bottom': c['bottom']}
                            for c in p.chars if c['text'].strip()])
            return out, 'text'
    if not AGING_RASTER_OK:
        raise ValueError("PDF นี้ฝังฟอนต์แบบพิเศษ (Type3) ต้องใช้ตัวอ่าน raster — "
                         "กรุณาติดตั้ง pypdfium2, numpy, Pillow")
    rot = _ag_rotation(pdf_bytes)
    pages = raster_words_pages_cached(pdf_bytes, rot)
    out = []
    for words in pages:
        chs = []
        for w in words:
            gl = w.get('glyphs') or []
            if not gl:
                chs.append({'text': w['text'], 'x0': w['x0'], 'x1': w['x1'],
                            'top': w['top'], 'bottom': w['bottom']})
                continue
            scale_x0 = w['x0']
            for g, ch in zip(gl, list(w['text'])):
                pass
            # glyph coords are absolute px/scale already relative to band
            txt = list(w['text'])
            for i, g in enumerate(gl):
                if i >= len(txt): break
                chs.append({'text': txt[i], 'x0': g['a']/10.0, 'x1': g['b']/10.0,
                            'top': w['top'], 'bottom': w['bottom']})
        out.append(chs)
    return out, 'raster'

def _ag_chars_to_words(chars, gap=0.9):
    chars = sorted(chars, key=lambda c: c['x0'])
    if not chars: return []
    words, cur = [], [chars[0]]
    for c in chars[1:]:
        if c['x0'] - cur[-1]['x1'] > gap: words.append(cur); cur = [c]
        else: cur.append(c)
    words.append(cur)
    return [{'text': ''.join(c['text'] for c in w), 'x0': w[0]['x0'], 'x1': w[-1]['x1'],
             'top': min(c['top'] for c in w), 'bottom': max(c['bottom'] for c in w)}
            for w in words]

def parse_ar_detailed(pdf_bytes):
    pages, source = _ag_get_chars_pages(pdf_bytes)
    meta = {'report': 'AR Detailed Aging', 'source': source, 'property': '',
            'print_date': '', 'print_time': '', 'business_date': '', 'age_credits': ''}
    items = []
    cur_account = None      # {'name','accno','rows':[],'total':None}
    cur_section = None
    last_kind = None

    def close_account():
        nonlocal cur_account
        if cur_account is not None:
            items.append(cur_account)
            cur_account = None

    for pi, chars in enumerate(pages):
        lines = _ag_lines(chars, tol=2.8)
        # header line (Guest Name Invoice No. ... Total)
        hdr_i = None; zones = None
        for li, line in enumerate(lines):
            ws = _ag_chars_to_words(line)
            texts = [w['text'] for w in ws]
            if 'Guest' in texts and 'Invoice' in texts and 'Post' in texts:
                def x0of(t):
                    return [w for w in ws if w['text'] == t][0]['x0']
                inv_x0 = x0of('Invoice'); folio_x0 = x0of('Folio'); post_x0 = x0of('Post')
                dates = [w for w in ws if w['text'] == 'Date']
                date_end = dates[-1]['x1'] + 8
                anchors = []
                ok = True
                for aw in ['30', '60', '90', '120', '150', 'Over', 'Total']:
                    cand = [w for w in ws if w['text'] == aw]
                    if not cand: ok = False; break
                    anchors.append(cand[-1]['x1'])
                if not ok: continue
                hdr_i = li
                zones = (inv_x0, folio_x0, post_x0, date_end, anchors)
                break
        if zones is None: continue
        inv_x0, folio_x0, post_x0, date_end, anchors = zones
        if pi == 0:
            for line in lines[:hdr_i]:
                ws = _ag_chars_to_words(line)
                raw = ' '.join(w['text'] for w in ws)
                if re.match(r'^\d{2}:\d{2}$', raw.strip()):
                    meta['print_time'] = raw.strip(); continue
                glue = _ag_glue(ws)
                if 'DETAILED' in glue or 'AGING' in glue: continue
                if 'ACCOUNTNAME' in glue: continue
                letters = sum(1 for ch in raw if ch.isalpha())
                if letters >= 8 and not meta['property']:
                    toks = [w['text'] for w in ws]
                    if re.match(r'^\d{2}/\d{2}/\d{2}$', toks[-1]):
                        meta['print_date'] = toks[-1]; toks = toks[:-1]
                    meta['property'] = ' '.join(toks)
        footer = False
        for li, line in enumerate(lines):
            if li <= hdr_i: continue
            ws = _ag_chars_to_words(line)
            glue = _ag_glue(ws)
            raw = ' '.join(w['text'] for w in ws)
            if footer or 'FILTER' in glue or 'SORTORDER' in glue or \
               'BUSINESSDATE' in glue or ('PAGE' in glue and 'OF' in glue):
                footer = True
                if 'BUSINESSDATE' in glue and not meta['business_date']:
                    m = re.search(r'(\d{2}/\d{2}/\d{2})', raw)
                    if m: meta['business_date'] = m.group(1)
                if 'AGECREDIT' in glue and not meta['age_credits']:
                    m2 = re.search(r'Credits?\s+([YN])\b', raw)
                    if m2: meta['age_credits'] = m2.group(1)
                continue
            if 'GUESTNAME' in glue or 'ACCOUNTNAME' in glue:
                continue
            # cells
            name_c   = [c for c in line if c['x1'] <= inv_x0 - 1.0]
            inv_c    = [c for c in line if c['x0'] >= inv_x0 - 1.5 and c['x0'] < folio_x0 - 1.5]
            folio_c  = [c for c in line if c['x0'] >= folio_x0 - 1.5 and c['x0'] < post_x0 - 1.5]
            date_c   = [c for c in line if c['x0'] >= post_x0 - 1.5 and c['x0'] < date_end]
            num_c    = [c for c in line if c['x0'] >= date_end]
            name_txt  = ' '.join(w['text'] for w in _ag_chars_to_words(name_c)).strip()
            inv_txt   = ''.join(c['text'] for c in sorted(inv_c, key=lambda c: c['x0'])).strip()
            folio_txt = ''.join(c['text'] for c in sorted(folio_c, key=lambda c: c['x0'])).strip()
            date_txt  = ''.join(c['text'] for c in sorted(date_c, key=lambda c: c['x0'])).strip()
            nwords = _ag_merge_line_numbers(_ag_chars_to_words(num_c))
            frag = defaultdict(list)
            for w in nwords:
                if not (re.match(r'^-?[\d.,]+$', w['text']) or w['text'] == '-'):
                    continue
                ds = [abs(w['x1'] - a) for a in anchors]
                bi = ds.index(min(ds))
                if ds[bi] <= 30:
                    frag[bi].append((w['x0'], w['text']))
            bucket_vals = [None]*7
            for k, lst in frag.items():
                bucket_vals[k] = _ag_amount(''.join(t for _, t in sorted(lst)))
            has_nums = any(v is not None for v in bucket_vals)

            if glue in ('A/RLEDGER', 'ARLEDGER') and not has_nums:
                if cur_section != 'A/R Ledger':
                    close_account()
                    cur_section = 'A/R Ledger'
                    items.append({'kind': 'section', 'label': 'A/R Ledger'})
                last_kind = 'section'; continue
            if re.match(r'^ACCR[UO]A[LI]S$', glue) and not has_nums:
                if cur_section != 'Accruals':
                    close_account()
                    cur_section = 'Accruals'
                    items.append({'kind': 'section', 'label': 'Accruals'})
                last_kind = 'section'; continue
            if re.match(r'^TOTALA/?RLEDGER$', glue):
                close_account()
                items.append({'kind': 'ledger_total', 'vals': bucket_vals})
                last_kind = 'ledger_total'; continue
            if re.match(r'^TOTALACCR[UO]A[LI]S$', glue):
                close_account()
                items.append({'kind': 'accruals_total', 'vals': bucket_vals})
                last_kind = 'accruals_total'; continue
            if glue == 'GRANDTOTAL':
                close_account()
                items.append({'kind': 'grand_total', 'vals': bucket_vals})
                last_kind = 'grand_total'; continue
            if glue == 'TOTAL' and has_nums:
                if cur_account is not None:
                    cur_account['total'] = bucket_vals
                    close_account()
                last_kind = 'account_total'; continue
            if has_nums:
                if cur_account is None:
                    cur_account = {'kind': 'account', 'name': '(unknown)', 'accno': '',
                                   'section': cur_section, 'rows': [], 'total': None}
                cur_account['rows'].append({'guest': name_txt, 'invoice': inv_txt,
                                            'folio': folio_txt, 'date': date_txt,
                                            'vals': bucket_vals})
                last_kind = 'row'; continue
            # no numbers
            accno_raw = (inv_txt + folio_txt).strip()
            is_accno = bool(re.match(r'^[A-Za-z0-9]{5,12}$', accno_raw)) and \
                       bool(re.search(r'\d{3}', accno_raw))
            if name_txt and (is_accno or last_kind in ('section', 'account_total',
                                                       None, 'ledger_total')):
                accno = _ag_fix_accno(accno_raw) if is_accno else ''
                if cur_account is not None and cur_account['rows'] and \
                   cur_account['name'] == name_txt and \
                   (cur_account['accno'] == accno or not accno or not cur_account['accno']):
                    if accno and not cur_account['accno']:
                        cur_account['accno'] = accno
                    last_kind = 'account_header'; continue   # repeated header on new page
                if cur_account is not None and not cur_account['rows']:
                    if name_txt.strip().lower() == cur_account['name'].strip().lower():
                        # same header repeated across a page break
                        if accno and not cur_account['accno']:
                            cur_account['accno'] = accno
                    else:
                        # consecutive headers -> treat as name continuation
                        cur_account['name'] = (cur_account['name'] + ' ' + name_txt).strip()
                        if accno: cur_account['accno'] = accno
                    last_kind = 'account_header'; continue
                close_account()
                cur_account = {'kind': 'account', 'name': name_txt, 'accno': accno,
                               'section': cur_section, 'rows': [], 'total': None}
                last_kind = 'account_header'; continue
            if name_txt and cur_account is not None and cur_account['rows'] and \
               last_kind in ('row', 'cont'):
                g = cur_account['rows'][-1]
                g['guest'] = (g['guest'] + ' ' + name_txt).strip()
                last_kind = 'cont'; continue
    close_account()
    if not any(it.get('kind') == 'account' and it.get('rows') for it in items):
        raise ValueError("ไม่พบข้อมูลใน PDF — กรุณาตรวจสอบว่าเป็น AR Detailed Aging ที่รองรับ")
    if meta.get('source') == 'raster':
        _ag_title_refine(pdf_bytes, meta)
    return meta, items

def validate_ar_detailed(items):
    issues, checks = [], 0
    ledger, accrual, grand = None, None, None
    sec_sums = {'A/R Ledger': [0.0]*7, 'Accruals': [0.0]*7, None: [0.0]*7}
    for it in items:
        k = it.get('kind')
        if k == 'account':
            for r in it['rows']:
                checks += 1
                b = r['vals']
                vals6 = [x or 0 for x in b[:6]]
                if b[6] is not None and abs(sum(vals6) - b[6]) > 0.015:
                    issues.append(f"{it['name']} / {r['guest'][:25]}: ผลรวม bucket ≠ Total")
            if it['total'] is not None:
                checks += 1
                for ci in range(7):
                    s = sum((r['vals'][ci] or 0) for r in it['rows'])
                    if abs(s - (it['total'][ci] or 0)) > 0.015:
                        issues.append(f"บัญชี {it['name']} คอลัมน์ {ci+1}: {s:,.2f} ≠ {(it['total'][ci] or 0):,.2f}")
                for ci in range(7):
                    sec_sums[it.get('section')][ci] += (it['total'][ci] or 0)
        elif k == 'ledger_total':
            checks += 1; ledger = it['vals']
            for ci in range(7):
                if abs(sec_sums['A/R Ledger'][ci] - (it['vals'][ci] or 0)) > 0.015:
                    issues.append(f"Total A/R Ledger คอลัมน์ {ci+1}: "
                                  f"{sec_sums['A/R Ledger'][ci]:,.2f} ≠ {(it['vals'][ci] or 0):,.2f}")
        elif k == 'accruals_total':
            checks += 1; accrual = it['vals']
            for ci in range(7):
                if abs(sec_sums['Accruals'][ci] - (it['vals'][ci] or 0)) > 0.015:
                    issues.append(f"Total Accruals คอลัมน์ {ci+1}: "
                                  f"{sec_sums['Accruals'][ci]:,.2f} ≠ {(it['vals'][ci] or 0):,.2f}")
        elif k == 'grand_total':
            checks += 1; grand = it['vals']
            if ledger is not None:
                for ci in range(7):
                    s = (ledger[ci] or 0) + ((accrual[ci] or 0) if accrual else 0)
                    if abs(s - (it['vals'][ci] or 0)) > 0.015:
                        issues.append(f"Grand Total คอลัมน์ {ci+1}: {s:,.2f} ≠ {(it['vals'][ci] or 0):,.2f}")
    return checks, issues

def _ag_title_refine(pdf_bytes, meta):
    """Optional: sharpen the italic title lines with tesseract when available.
    Purely cosmetic (header cells); numeric data never depends on this."""
    try:
        import shutil
        if not shutil.which('tesseract'): return
        import pytesseract
        doc = _pdfium.PdfDocument(io.BytesIO(pdf_bytes))
        pil = doc[0].render(scale=8).to_pil()
        if _ag_rotation(pdf_bytes): pil = pil.transpose(_PILImage.ROTATE_270)
        strip = pil.crop((0, 0, pil.width, min(int(60*8), pil.height)))
        txt = pytesseract.image_to_string(strip, config='--psm 6')
        for l in (l.strip() for l in txt.splitlines()):
            if not l: continue
            m = re.match(r'^(.*?)\s+(\d{2}/\d{2}/\d{2})$', l)
            if m and sum(ch.isalpha() for ch in m.group(1)) >= 8:
                prop = re.sub(r'^[^A-Za-z0-9]+', '', m.group(1)).strip()
                if prop: meta['property'] = prop
                meta['print_date'] = m.group(2)
                continue
            m = re.match(r'^(\d{2}:\d{2})$', l)
            if m: meta['print_time'] = m.group(1)
    except Exception:
        pass

# ─── Excel writers ───────────────────────────────────────────────────
_AG_NUM_FMT = '#,##0.00;[Red]- #,##0.00;0.00'
_AG_PCT_FMT = '0.00" %";[Red]-0.00" %";0.00" %"'

def _ag_styles():
    thin   = Side(style='thin',   color='BFBFBF')
    medium = Side(style='medium', color='2E75B6')
    return {
        'white_f':  Font(name='Arial', bold=True, size=9, color="FFFFFF"),
        'data_f':   Font(name='Arial', size=9),
        'bold_f':   Font(name='Arial', bold=True, size=9),
        'ital_f':   Font(name='Arial', italic=True, size=9, color="595959"),
        'title_f':  Font(name='Arial', bold=True, size=13, color="1F4E79"),
        'sub_f':    Font(name='Arial', bold=True, size=11, color="2E75B6"),
        'small_b':  Font(name='Arial', bold=True, size=9),
        'small':    Font(name='Arial', size=9),
        'c':  Alignment(horizontal='center', vertical='center', wrap_text=True),
        'l':  Alignment(horizontal='left',   vertical='center', wrap_text=True),
        'r':  Alignment(horizontal='right',  vertical='center'),
        'thin': thin, 'medium': medium,
        'tb':  Border(left=thin, right=thin, top=thin, bottom=thin),
        'mb':  Border(left=thin, right=thin, top=medium, bottom=medium),
        'hdr_fill': PatternFill("solid", fgColor="1F4E79"),
        'sub_fill': PatternFill("solid", fgColor="2E75B6"),
        'tot_fill': PatternFill("solid", fgColor="D6E4F0"),
        'sec_fill': PatternFill("solid", fgColor="EAF0FB"),
        'alt_fill': PatternFill("solid", fgColor="F2F7FB"),
        'wht_fill': PatternFill("solid", fgColor="FFFFFF"),
        'grand_fill': PatternFill("solid", fgColor="BDD7EE"),
    }

_AG_BUCKET_HEADERS = ['Up to 30', '31 - 60', '61 - 90', '91 - 120',
                      '121 - 150', '151 and Over', 'Total']

def convert_aging_summary(pdf_bytes):
    meta, items = parse_aging_summary(pdf_bytes)
    checks, issues = validate_aging_summary(items)
    st_ = _ag_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aging Summary"
    NC = 10   # A..J

    ws.merge_cells('A1:J1')
    ws['A1'] = meta['property'] or 'Aging Summary'
    ws['A1'].font = st_['title_f']; ws['A1'].alignment = st_['c']
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:G2')
    ws['A2'] = 'Aging Summary for All Types'
    ws['A2'].font = st_['sub_f']; ws['A2'].alignment = st_['l']
    ws['H2'] = 'Print Date:'; ws['H2'].font = st_['small_b']; ws['H2'].alignment = st_['r']
    ws.merge_cells('I2:J2')
    ws['I2'] = (meta['print_date'] + ' ' + meta['print_time']).strip()
    ws['I2'].font = st_['small']; ws['I2'].alignment = st_['c']

    ws['A3'] = f"Age Credits : {meta['age_credits']}".strip()
    ws['A3'].font = st_['small']
    ws['H3'] = 'Business Date:'; ws['H3'].font = st_['small_b']; ws['H3'].alignment = st_['r']
    ws.merge_cells('I3:J3')
    ws['I3'] = meta['business_date']; ws['I3'].font = st_['small']; ws['I3'].alignment = st_['c']
    ws.row_dimensions[4].height = 5

    headers = ['Account Name', 'Account No.', 'Credit Limit / LOC'] + _AG_BUCKET_HEADERS
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = st_['white_f']; c.fill = st_['hdr_fill']
        c.alignment = st_['c']; c.border = st_['tb']
    ws.row_dimensions[5].height = 24

    er = 6
    acc_i = 0
    for it in items:
        k = it['kind']
        if k == 'type_header':
            ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=3)
            c = ws.cell(row=er, column=1,
                        value=f"Account Type {it['code']}   {it['desc']}".strip())
            c.font = Font(name='Arial', bold=True, size=9, color="FFFFFF")
            c.alignment = st_['l']
            for col in range(1, NC+1):
                ws.cell(row=er, column=col).fill = st_['sub_fill']
                ws.cell(row=er, column=col).border = st_['tb']
            ws.row_dimensions[er].height = 16
            acc_i = 0
        elif k == 'account':
            fill = st_['alt_fill'] if acc_i % 2 == 0 else st_['wht_fill']
            vals = [it['name'], it['accno'], it['climit']] + it['vals']
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=er, column=col, value=v)
                c.font = st_['data_f']; c.fill = fill; c.border = st_['tb']
                if col == 1: c.alignment = st_['l']
                elif col == 2: c.alignment = st_['c']
                else:
                    c.alignment = st_['r']
                    if v is not None: c.number_format = _AG_NUM_FMT
            ws.row_dimensions[er].height = 15
            acc_i += 1
        elif k in ('type_total', 'ledger_total', 'accruals_total', 'grand_total'):
            label = {'type_total': f"Total for Account Type {it.get('code','')}",
                     'ledger_total': 'Total AR Ledger',
                     'accruals_total': 'Total Accruals',
                     'grand_total': 'Grand Total'}[k]
            fill = st_['grand_fill'] if k == 'grand_total' else st_['tot_fill']
            ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=2)
            c = ws.cell(row=er, column=1, value=label)
            c.font = st_['bold_f']; c.alignment = st_['l']
            cnt = it.get('count')
            if cnt is not None:
                c3 = ws.cell(row=er, column=3, value=cnt)
                c3.alignment = st_['r']; c3.number_format = '0'
            for col, v in enumerate(it['vals'], 4):
                c = ws.cell(row=er, column=col, value=v)
                c.alignment = st_['r']
                if v is not None: c.number_format = _AG_NUM_FMT
            for col in range(1, NC+1):
                cc = ws.cell(row=er, column=col)
                cc.fill = fill; cc.border = st_['mb']
                cc.font = st_['bold_f']
            ws.row_dimensions[er].height = 17
        elif k == 'pct':
            ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=2)
            c = ws.cell(row=er, column=1, value='No of Accounts & Percentage')
            c.font = st_['ital_f']; c.alignment = st_['l']
            if it.get('count') is not None:
                c3 = ws.cell(row=er, column=3, value=it['count'])
                c3.alignment = st_['r']; c3.number_format = '0'
                c3.font = st_['ital_f']
            for col, v in enumerate(it['vals'], 4):
                c = ws.cell(row=er, column=col, value=v)
                c.alignment = st_['r']; c.font = st_['ital_f']
                if v is not None: c.number_format = _AG_PCT_FMT
            for col in range(1, NC+1):
                ws.cell(row=er, column=col).border = st_['tb']
            ws.row_dimensions[er].height = 14
        elif k == 'section':
            ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=NC)
            c = ws.cell(row=er, column=1, value=it['label'])
            c.font = Font(name='Arial', bold=True, size=10, color="1F4E79")
            c.alignment = st_['l']; 
            for col in range(1, NC+1):
                ws.cell(row=er, column=col).fill = st_['sec_fill']
                ws.cell(row=er, column=col).border = st_['tb']
            ws.row_dimensions[er].height = 17
        er += 1

    widths = [36, 12, 15, 13, 13, 13, 13, 13, 13, 14]
    for i, cw in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = cw
    ws.freeze_panes = 'A6'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    n_accounts = sum(1 for it in items if it['kind'] == 'account')
    grand = next((it['vals'] for it in items if it['kind'] == 'grand_total'), [None]*7)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, n_accounts, grand, checks, issues, meta

def convert_ar_detailed(pdf_bytes):
    meta, items = parse_ar_detailed(pdf_bytes)
    checks, issues = validate_ar_detailed(items)
    st_ = _ag_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR Detailed Aging"
    NC = 13   # A..M

    ws.merge_cells('A1:M1')
    ws['A1'] = meta['property'] or 'AR Detailed Aging'
    ws['A1'].font = st_['title_f']; ws['A1'].alignment = st_['c']
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:J2')
    ws['A2'] = 'AR Detailed Aging'
    ws['A2'].font = st_['sub_f']; ws['A2'].alignment = st_['l']
    ws['K2'] = 'Print Date:'; ws['K2'].font = st_['small_b']; ws['K2'].alignment = st_['r']
    ws.merge_cells('L2:M2')
    ws['L2'] = (meta['print_date'] + ' ' + meta['print_time']).strip()
    ws['L2'].font = st_['small']; ws['L2'].alignment = st_['c']

    ws['A3'] = f"Age Credits : {meta['age_credits']}".strip()
    ws['A3'].font = st_['small']
    ws['K3'] = 'Business Date:'; ws['K3'].font = st_['small_b']; ws['K3'].alignment = st_['r']
    ws.merge_cells('L3:M3')
    ws['L3'] = meta['business_date']; ws['L3'].font = st_['small']; ws['L3'].alignment = st_['c']
    ws.row_dimensions[4].height = 5

    headers = ['Account Name', 'Account No.', 'Guest Name', 'Invoice No.', 'Folio No.',
               'Post Date'] + _AG_BUCKET_HEADERS
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = st_['white_f']; c.fill = st_['hdr_fill']
        c.alignment = st_['c']; c.border = st_['tb']
    ws.row_dimensions[5].height = 24

    er = 6
    n_rows = 0
    for it in items:
        k = it.get('kind')
        if k == 'section':
            ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=NC)
            c = ws.cell(row=er, column=1, value=it['label'])
            c.font = Font(name='Arial', bold=True, size=10, color="1F4E79")
            c.alignment = st_['l']
            for col in range(1, NC+1):
                ws.cell(row=er, column=col).fill = st_['sec_fill']
                ws.cell(row=er, column=col).border = st_['tb']
            ws.row_dimensions[er].height = 17
            er += 1
        elif k == 'account':
            c = ws.cell(row=er, column=1, value=it['name'])
            c.font = st_['bold_f']; c.alignment = st_['l']
            c2 = ws.cell(row=er, column=2, value=it['accno'])
            c2.font = st_['bold_f']; c2.alignment = st_['c']
            for col in range(1, NC+1):
                ws.cell(row=er, column=col).fill = st_['sec_fill']
                ws.cell(row=er, column=col).border = st_['tb']
            ws.row_dimensions[er].height = 15
            er += 1
            for i, r in enumerate(it['rows']):
                fill = st_['alt_fill'] if i % 2 == 0 else st_['wht_fill']
                inv = int(r['invoice']) if r['invoice'].isdigit() else (r['invoice'] or None)
                fol = int(r['folio']) if r['folio'].isdigit() else (r['folio'] or None)
                vals = [None, None, r['guest'], inv, fol, r['date'] or None] + r['vals']
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=er, column=col, value=v)
                    c.font = st_['data_f']; c.fill = fill; c.border = st_['tb']
                    if col == 3: c.alignment = st_['l']
                    elif col in (4, 5, 6): c.alignment = st_['c']
                    elif col >= 7:
                        c.alignment = st_['r']
                        if v is not None: c.number_format = _AG_NUM_FMT
                ws.row_dimensions[er].height = 14
                n_rows += 1
                er += 1
            if it['total'] is not None:
                ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=6)
                c = ws.cell(row=er, column=1, value=f"Total   {it['name']}")
                c.font = st_['bold_f']; c.alignment = st_['l']
                for col, v in enumerate(it['total'], 7):
                    cc = ws.cell(row=er, column=col, value=v)
                    cc.alignment = st_['r']
                    if v is not None: cc.number_format = _AG_NUM_FMT
                for col in range(1, NC+1):
                    cc = ws.cell(row=er, column=col)
                    cc.fill = st_['tot_fill']; cc.border = st_['tb']
                    cc.font = st_['bold_f']
                ws.row_dimensions[er].height = 15
                er += 1
        elif k in ('ledger_total', 'accruals_total', 'grand_total'):
            label = {'ledger_total': 'Total A/R Ledger',
                     'accruals_total': 'Total Accruals',
                     'grand_total': 'Grand Total'}[k]
            fill = st_['grand_fill'] if k == 'grand_total' else st_['tot_fill']
            ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=6)
            c = ws.cell(row=er, column=1, value=label)
            c.font = st_['bold_f']; c.alignment = st_['l']
            for col, v in enumerate(it['vals'], 7):
                cc = ws.cell(row=er, column=col, value=v)
                cc.alignment = st_['r']
                if v is not None: cc.number_format = _AG_NUM_FMT
            for col in range(1, NC+1):
                cc = ws.cell(row=er, column=col)
                cc.fill = fill; cc.border = st_['mb']; cc.font = st_['bold_f']
            ws.row_dimensions[er].height = 17
            er += 1

    widths = [30, 11, 30, 10, 10, 10, 12, 12, 12, 12, 12, 12, 13]
    for i, cw in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = cw
    ws.freeze_panes = 'A6'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    grand = next((it['vals'] for it in items if it.get('kind') == 'grand_total'), [None]*7)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, n_rows, grand, checks, issues, meta

# ─── detection ───────────────────────────────────────────────────────
_AG_SNIFF_CACHE = {}
def detect_aging_type(pdf_bytes, progress_cb=None):
    """'aging_summary' | 'ar_detailed_aging' | None"""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text0 = pdf.pages[0].extract_text() or ''
    except Exception:
        return None
    if not _ag_is_garbled(text0):
        up = text0.upper()
        if 'AR DETAILED AGING' in up: return 'ar_detailed_aging'
        if 'AGING SUMMARY' in up: return 'aging_summary'
        return None
    if not AGING_RASTER_OK:
        return None
    import hashlib
    key = hashlib.md5(pdf_bytes).hexdigest()
    if key in _AG_SNIFF_CACHE: return _AG_SNIFF_CACHE[key]
    rot = _ag_rotation(pdf_bytes)
    words = raster_words_pages_cached(pdf_bytes, rot, progress_cb)[0]
    glue = ''.join(w['text'] for w in words).upper().replace('1', 'I').replace('0', 'O')
    glue = re.sub(r'[^A-Z/&]', '', glue)
    result = None
    if 'GUESTNAME' in glue and 'POSTDATE' in glue:
        result = 'ar_detailed_aging'
    elif 'CREDITLIMIT' in glue and ('UPTO' in glue or 'ANDOVER' in glue):
        result = 'aging_summary'
    _AG_SNIFF_CACHE[key] = result
    return result

_AG_RASTER_CACHE = {}
def raster_words_pages_cached(pdf_bytes, rotate, progress_cb=None):
    import hashlib
    key = (hashlib.md5(pdf_bytes).hexdigest(), rotate)
    if key not in _AG_RASTER_CACHE:
        if len(_AG_RASTER_CACHE) > 3: _AG_RASTER_CACHE.clear()
        _AG_RASTER_CACHE[key] = raster_words_pages(pdf_bytes, rotate=rotate,
                                                   progress_cb=progress_cb)
    return _AG_RASTER_CACHE[key]


def detect_aging_type_safe(pdf_bytes, progress_cb=None):
    try:
        return detect_aging_type(pdf_bytes, progress_cb)
    except Exception:
        return None


# ─── UI ─────────────────────────────────────────────────────────────
st.markdown("## 📊 PDF → Excel ตัวไหนไม่ได้แจ้งก้องนะครับ")
st.markdown("รองรับ 5 ประเภท: 1.งบทดลอง 2.Statement of Account 3.Matrix trial balance 4.AR Detailed Aging 5.Aging Summary")
st.divider()

uploaded = st.file_uploader(
    "อัปโหลดไฟล์ PDF",
    type=['pdf'],
    help="รองรับ PDF ที่ไม่ใช่ภาพสแกน"
)

if uploaded:
    pdf_bytes = uploaded.read()

    _prog = st.empty()
    def _prog_cb(i, n):
        _prog.progress(i / n, text=f"กำลังถอดรหัสฟอนต์ในไฟล์ — หน้า {i}/{n}")
    aging_type = detect_aging_type_safe(pdf_bytes, progress_cb=_prog_cb)
    _prog.empty()

    pdf_type = aging_type if aging_type else detect_pdf_type(pdf_bytes)

    def fmt(n): return f"{n:,.2f}" if n else "-"

    def _aging_ui(kind):
        conv  = convert_aging_summary if kind == 'aging_summary' else convert_ar_detailed
        label = 'Aging Summary' if kind == 'aging_summary' else 'AR Detailed Aging'
        unit  = 'บัญชี' if kind == 'aging_summary' else 'รายการ'
        st.info(f"📅 ตรวจพบ: **{label}**")

        if st.button("🔄 แปลงเป็น Excel", type="primary", use_container_width=True):
            with st.spinner("กำลังอ่าน PDF และสร้างไฟล์ Excel..."):
                try:
                    buf, n, grand, checks, issues, meta = conv(pdf_bytes)
                    xlsx_name = uploaded.name.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')

                    st.success(f"✅ แปลงสำเร็จ — **{n:,} {unit}** ({meta.get('property','')})")

                    col1, col2, col3 = st.columns(3)
                    col1.metric(f"จำนวน{unit}", f"{n:,}")
                    col2.metric("Grand Total (Up to 30)", fmt(grand[0] if grand else None))
                    col3.metric("Grand Total (Total)",    fmt(grand[6] if grand else None))

                    if issues:
                        st.warning("⚠️ พบข้อสังเกตจากการตรวจทานยอด "
                                   f"{len(issues)} จุด (จากทั้งหมด {checks} จุด)")
                        with st.expander("รายละเอียดการตรวจทาน"):
                            for it in issues:
                                st.markdown(f"- {it}")
                    else:
                        st.caption(f"🔎 ตรวจทานผลรวมอัตโนมัติ {checks} จุด — ตรงกับ PDF ทั้งหมด")

                    st.download_button(
                        label="⬇️ ดาวน์โหลด Excel",
                        data=buf, file_name=xlsx_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary"
                    )
                except Exception as e:
                    st.error(f"❌ เกิดข้อผิดพลาด: {e}")

    if pdf_type in ('aging_summary', 'ar_detailed_aging'):
        _aging_ui(pdf_type)

    elif pdf_type == 'statement':
        client_name = detect_company_soa(pdf_bytes)
        print_date  = get_print_date(pdf_bytes)
        st.info(f"📄 ตรวจพบ: **Statement of Account** — {client_name}")

        if st.button("🔄 แปลงเป็น Excel", type="primary", use_container_width=True):
            with st.spinner("กำลังอ่าน PDF และสร้างไฟล์ Excel..."):
                try:
                    buf, row_count, td, tc, tb = convert_statement(pdf_bytes, client_name, print_date)
                    xlsx_name = uploaded.name.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')

                    st.success(f"✅ แปลงสำเร็จ — **{row_count:,} รายการ**")

                    col1, col2, col3 = st.columns(3)
                    col1.metric("Total Debit",   fmt(td))
                    col2.metric("Total Credit",  fmt(tc) if tc else "-")
                    col3.metric("Total Balance", fmt(tb))

                    st.download_button(
                        label="⬇️ ดาวน์โหลด Excel",
                        data=buf, file_name=xlsx_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary"
                    )
                except Exception as e:
                    st.error(f"❌ เกิดข้อผิดพลาด: {e}")

    elif pdf_type == 'matrix_trial_balance':
        company_name, report_date = detect_info_matrix(pdf_bytes)
        st.info(f"📊 ตรวจพบ: **New Matrix Trial Balance** — {company_name}")

        if st.button("🔄 แปลงเป็น Excel", type="primary", use_container_width=True):
            with st.spinner("กำลังอ่าน PDF และสร้างไฟล์ Excel..."):
                try:
                    buf, row_count, total_net = convert_matrix_trial_balance(
                        pdf_bytes, company_name, report_date)
                    xlsx_name = uploaded.name.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')

                    st.success(f"✅ แปลงสำเร็จ — **{row_count:,} รายการ**")

                    col1, col2 = st.columns(2)
                    col1.metric("รายการทั้งหมด", f"{row_count:,}")
                    col2.metric("Net Amount รวม", fmt(total_net))

                    st.download_button(
                        label="⬇️ ดาวน์โหลด Excel",
                        data=buf, file_name=xlsx_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary"
                    )
                except Exception as e:
                    st.error(f"❌ เกิดข้อผิดพลาด: {e}")

    else:
        company_name = detect_company_tb(pdf_bytes)
        st.info(f"🏢 ตรวจพบ: **งบทดลอง** — {company_name}")

        if st.button("🔄 แปลงเป็น Excel", type="primary", use_container_width=True):
            with st.spinner("กำลังอ่านตาราง PDF และสร้างไฟล์ Excel..."):
                try:
                    buf, row_count, totals = convert_trial_balance(pdf_bytes, company_name)
                    xlsx_name = uploaded.name.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')

                    st.success(f"✅ แปลงสำเร็จ — **{row_count:,} รายการ**")

                    col1, col2, col3 = st.columns(3)
                    col1.metric("ยอดยกมา (Dr)",   fmt(totals[0]))
                    col2.metric("ยอดสะสม (Dr)",   fmt(totals[2]))
                    col3.metric("ยอดยกไป (Dr)",   fmt(totals[4]))

                    st.download_button(
                        label="⬇️ ดาวน์โหลด Excel",
                        data=buf, file_name=xlsx_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary"
                    )
                except Exception as e:
                    st.error(f"❌ เกิดข้อผิดพลาด: {e}")

st.divider()
st.caption("💡 ระบบตรวจจับประเภท PDF อัตโนมัติ — รองรับเฉพาะ searchable PDF (ไม่ใช่ภาพสแกน)")
